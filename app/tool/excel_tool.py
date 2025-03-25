import os
import re
import logging
import openpyxl
from typing import Any, Dict, List, Optional, Union, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from app.tool.base import BaseTool, ToolResult

logger = logging.getLogger(__name__)

class ExcelTool(BaseTool):
    """A tool for manipulating Excel spreadsheets."""
    name: str = "excel_tool"
    description: str = "Manipulate Excel spreadsheets including reading, writing, and formatting operations"
    
    # Current workbook and worksheet state
    _workbook: Optional[Workbook] = None
    _current_sheet: Optional[Worksheet] = None
    _current_file_path: Optional[str] = None
    
    def __init__(self):
        """Initialize the Excel tool."""
        super().__init__(
            name="excel_tool",
            description="Manipulate Excel spreadsheets including reading, writing, and formatting operations"
        )
        
        # Define parameters schema based on all available functions
        self.parameters = {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "description": "The Excel action to perform",
                    "enum": [
                        "open_sheet", "del_sheet", "freeze_data", "get_A1_annotation",
                        "insert_cols", "insert_rows", "delete_batch_data", "update_cell",
                        "update_cell_by_formula", "update_range", "sort_sheet_by_col",
                        "merge_cells", "update_note", "get_all_values", "get_range_values",
                        "get_cell_value", "get_value_by_formula", "filter_cells", "get_note",
                        "finish", "create_workbook", "save_workbook", "open_workbook",
                        # Add common action aliases
                        "read", "read_sheet", "open", "save", "create", "write", "update",
                        "get", "get_range", "insert_row", "insert_column", "delete", "sort",
                        "filter", "merge"
                    ]
                },
                "params": {
                    "type": "object",
                    "description": "Parameters specific to the chosen action"
                }
            },
            "required": ["action", "params"]
        }
        
    async def execute(self, action: str, params: Dict[str, Any]) -> Dict[str, Any]:
        """Execute the specified Excel action with the given parameters.
        
        Args:
            action: The Excel action to perform.
            params: Parameters specific to the action.
            
        Returns:
            The result of the Excel operation.
        """
        # Map common action names to actual Excel tool actions
        action_map = {
            "read": "get_all_values",
            "read_sheet": "get_all_values",
            "open": "open_workbook",
            "save": "save_workbook",
            "create": "create_workbook",
            "write": "update_cell",
            "update": "update_cell",
            "get": "get_cell_value",
            "get_range": "get_range_values",
            "insert_row": "insert_rows",
            "insert_column": "insert_cols",
            "delete": "delete_batch_data",
            "sort": "sort_sheet_by_col",
            "filter": "filter_cells",
            "merge": "merge_cells"
        }
        
        # Map the action if it's a common name
        mapped_action = action_map.get(action, action)
        
        # Log the action mapping if it occurred
        if mapped_action != action:
            logger.info(f"Mapped Excel action '{action}' to '{mapped_action}'")
            action = mapped_action
        
        # Special case handling for read/get_all_values when no workbook is open yet
        if action == "get_all_values" and (self._workbook is None) and "file_path" in params:
            logger.info(f"Auto-opening workbook for get_all_values: {params['file_path']}")
            # First open the workbook
            open_result = await self._open_workbook(file_path=params["file_path"])
            if "error" in open_result and open_result["error"]:
                return open_result
            
            # If sheet name is provided, open that sheet
            if "sheet_name" in params:
                sheet_result = await self._open_sheet(name=params["sheet_name"])
                if "error" in sheet_result and sheet_result["error"]:
                    return sheet_result
            
            # Now get all values without the file_path param
            params_copy = params.copy()
            params_copy.pop("file_path", None)
            return await self._get_all_values(**params_copy)
        
        try:
            if action == "open_workbook":
                return await self._open_workbook(**params)
            elif action == "create_workbook":
                return await self._create_workbook(**params)
            elif action == "save_workbook":
                return await self._save_workbook(**params)
            elif action == "open_sheet":
                return await self._open_sheet(**params)
            elif action == "del_sheet":
                return await self._del_sheet(**params)
            elif action == "freeze_data":
                return await self._freeze_data(**params)
            elif action == "get_A1_annotation":
                return await self._get_A1_annotation(**params)
            elif action == "insert_cols":
                return await self._insert_cols(**params)
            elif action == "insert_rows":
                return await self._insert_rows(**params)
            elif action == "delete_batch_data":
                return await self._delete_batch_data(**params)
            elif action == "update_cell":
                return await self._update_cell(**params)
            elif action == "update_cell_by_formula":
                return await self._update_cell_by_formula(**params)
            elif action == "update_range":
                return await self._update_range(**params)
            elif action == "sort_sheet_by_col":
                return await self._sort_sheet_by_col(**params)
            elif action == "merge_cells":
                return await self._merge_cells(**params)
            elif action == "update_note":
                return await self._update_note(**params)
            elif action == "get_all_values":
                return await self._get_all_values(**params)
            elif action == "get_range_values":
                return await self._get_range_values(**params)
            elif action == "get_cell_value":
                return await self._get_cell_value(**params)
            elif action == "get_value_by_formula":
                return await self._get_value_by_formula(**params)
            elif action == "filter_cells":
                return await self._filter_cells(**params)
            elif action == "get_note":
                return await self._get_note(**params)
            elif action == "finish":
                return await self._finish(**params)
            else:
                return ToolResult(error=f"Unknown action: {action}").dict()
        except Exception as e:
            logger.exception(f"Error executing Excel action {action}: {e}")
            return ToolResult(error=f"Error executing {action}: {str(e)}").dict()

    # Helper methods for cell reference conversions
    def _position_to_coordinates(self, position: str) -> Tuple[int, int]:
        """Convert A1 cell reference to row, column coordinates."""
        match = re.match(r'^([A-Z]+)(\d+)$', position)
        if not match:
            raise ValueError(f"Invalid cell position: {position}")
        
        col = column_index_from_string(match.group(1))
        row = int(match.group(2))
        return row, col
    
    def _coordinates_to_position(self, row: int, col: int) -> str:
        """Convert row, column coordinates to A1 cell reference."""
        return f"{get_column_letter(col)}{row}"
    
    def _ensure_workbook_open(self):
        """Ensure a workbook is open."""
        if self._workbook is None:
            raise ValueError("No workbook is currently open. Use open_workbook or create_workbook first.")
    
    def _ensure_sheet_selected(self):
        """Ensure a worksheet is selected."""
        self._ensure_workbook_open()
        if self._current_sheet is None:
            raise ValueError("No worksheet is currently selected. Use open_sheet first.")

    # Workbook operations
    async def _create_workbook(self, filename: str) -> Dict[str, Any]:
        """Create a new Excel workbook.
        
        Args:
            filename: The name of the file to create.
            
        Returns:
            Result of the operation.
        """
        try:
            self._workbook = Workbook()
            self._current_sheet = self._workbook.active
            self._current_file_path = filename
            return ToolResult(output=f"Created new workbook: {filename}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to create workbook: {str(e)}").dict()
    
    async def _open_workbook(self, filename: str = None, file_path: str = None) -> Dict[str, Any]:
        """Open an existing Excel workbook.
        
        Args:
            filename: The name of the file to open.
            file_path: Alternative parameter name for the file path to open.
            
        Returns:
            Result of the operation.
        """
        try:
            # Use file_path if provided, otherwise use filename
            filepath = file_path if file_path is not None else filename
            
            if not filepath:
                return ToolResult(error="No filename or file_path provided").dict()
                
            # Log the file path being opened
            logger.info(f"Opening Excel workbook: {filepath}")
            
            # Check if file exists
            if not os.path.exists(filepath):
                return ToolResult(error=f"File not found: {filepath}").dict()
                
            self._workbook = load_workbook(filepath)
            self._current_sheet = self._workbook.active
            self._current_file_path = filepath
            
            sheet_names = self._workbook.sheetnames
            active_sheet = self._current_sheet.title
            
            return ToolResult(
                output={
                    "message": f"Opened workbook {filepath}",
                    "sheet_names": sheet_names,
                    "active_sheet": active_sheet
                }
            ).dict()
        except Exception as e:
            logger.exception(f"Error opening workbook: {e}")
            return ToolResult(error=f"Error opening workbook: {str(e)}").dict()
    
    async def _save_workbook(self, filename: Optional[str] = None) -> Dict[str, Any]:
        """Save the current workbook.
        
        Args:
            filename: Optional new filename to save as.
            
        Returns:
            Result of the operation.
        """
        self._ensure_workbook_open()
        
        try:
            save_path = filename or self._current_file_path
            if not save_path:
                return ToolResult(error="No filename specified and no current file path").dict()
            
            self._workbook.save(save_path)
            if filename:
                self._current_file_path = filename
            
            return ToolResult(output=f"Workbook saved to: {save_path}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to save workbook: {str(e)}").dict()

    # Sheet operations
    async def _open_sheet(self, name: str) -> Dict[str, Any]:
        """Open a sheet by name.
        
        Args:
            name: The name of the sheet to open.
            
        Returns:
            Result of the operation.
        """
        self._ensure_workbook_open()
        
        try:
            if name not in self._workbook.sheetnames:
                return ToolResult(error=f"Sheet not found: {name}").dict()
            
            self._current_sheet = self._workbook[name]
            return ToolResult(output=f"Opened sheet: {name}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to open sheet: {str(e)}").dict()
    
    async def _del_sheet(self, name: str) -> Dict[str, Any]:
        """Delete a sheet by name.
        
        Args:
            name: The name of the sheet to delete.
            
        Returns:
            Result of the operation.
        """
        self._ensure_workbook_open()
        
        try:
            if name not in self._workbook.sheetnames:
                return ToolResult(error=f"Sheet not found: {name}").dict()
            
            # Can't delete the only remaining sheet
            if len(self._workbook.sheetnames) <= 1:
                return ToolResult(error="Cannot delete the only sheet in workbook").dict()
            
            # If deleting the current sheet, switch to another sheet
            if self._current_sheet.title == name:
                other_sheet = next(sheet for sheet in self._workbook.sheetnames if sheet != name)
                self._current_sheet = self._workbook[other_sheet]
            
            del self._workbook[name]
            return ToolResult(output=f"Deleted sheet: {name}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to delete sheet: {str(e)}").dict()
    
    async def _freeze_data(self, dimension: str, num: int) -> Dict[str, Any]:
        """Freeze rows and/or columns on the worksheet.
        
        Args:
            dimension: The dimension to freeze, either 'rows' or 'columns'.
            num: Number of rows/cols to freeze.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            if dimension.lower() == 'rows':
                self._current_sheet.freeze_panes = f"A{num+1}"
            elif dimension.lower() == 'columns':
                col_letter = get_column_letter(num+1)
                self._current_sheet.freeze_panes = f"{col_letter}1"
            else:
                return ToolResult(error=f"Invalid dimension: {dimension}. Must be 'rows' or 'columns'").dict()
            
            return ToolResult(output=f"Froze {dimension}: {num}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to freeze {dimension}: {str(e)}").dict()
    
    async def _get_A1_annotation(self, row: int, col: int) -> Dict[str, Any]:
        """Translate the cell position (row,col) into A1 annotation.
        
        Args:
            row: Row index.
            col: Column index.
            
        Returns:
            The A1 notation of the cell or an error message.
        """
        try:
            if row < 1 or col < 1:
                return ToolResult(error="Row and column indices must be positive").dict()
            
            position = self._coordinates_to_position(row, col)
            return ToolResult(output=position).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to get A1 annotation: {str(e)}").dict()
    
    # Data operations
    async def _insert_cols(self, values_list: List[List[Any]], col_idx: int = 1) -> Dict[str, Any]:
        """Insert columns into sheet at specified column index.
        
        Args:
            values_list: A list of lists, each list containing one column's values.
            col_idx: Start column to update. Defaults to 1.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            # Insert columns
            self._current_sheet.insert_cols(col_idx, len(values_list))
            
            # Fill with values
            for i, col_values in enumerate(values_list):
                col = col_idx + i
                for j, value in enumerate(col_values):
                    row = j + 1
                    self._current_sheet.cell(row=row, column=col, value=value)
            
            return ToolResult(output=f"Inserted {len(values_list)} columns at column {col_idx}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to insert columns: {str(e)}").dict()
    
    async def _insert_rows(self, values_list: List[List[Any]], row_idx: int = 1) -> Dict[str, Any]:
        """Insert rows into sheet at specified row index.
        
        Args:
            values_list: A list of lists, each list containing one row's values.
            row_idx: Start row to update. Defaults to 1.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            # Insert rows
            self._current_sheet.insert_rows(row_idx, len(values_list))
            
            # Fill with values
            for i, row_values in enumerate(values_list):
                row = row_idx + i
                for j, value in enumerate(row_values):
                    col = j + 1
                    self._current_sheet.cell(row=row, column=col, value=value)
            
            return ToolResult(output=f"Inserted {len(values_list)} rows at row {row_idx}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to insert rows: {str(e)}").dict()
    
    async def _delete_batch_data(self, dimension: str, index_list: List[int]) -> Dict[str, Any]:
        """Delete a batch of data in the sheet.
        
        Args:
            dimension: The dimension to delete, either 'row' or 'col'.
            index_list: List of the indexes of rows/cols for deletion.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            if dimension.lower() not in ['row', 'col']:
                return ToolResult(error=f"Invalid dimension: {dimension}. Must be 'row' or 'col'").dict()
            
            # Sort indices in reverse order to avoid shifting issues
            sorted_indices = sorted(index_list, reverse=True)
            
            for idx in sorted_indices:
                if dimension.lower() == 'row':
                    self._current_sheet.delete_rows(idx)
                else:  # col
                    self._current_sheet.delete_cols(idx)
            
            return ToolResult(output=f"Deleted {len(index_list)} {dimension}s").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to delete batch data: {str(e)}").dict()
    
    async def _update_cell(self, position: str, value: Any) -> Dict[str, Any]:
        """Update the value of a cell.
        
        Args:
            position: A1 notation of the cell position.
            value: The value to set.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            row, col = self._position_to_coordinates(position)
            self._current_sheet.cell(row=row, column=col, value=value)
            
            return ToolResult(output=f"Updated cell {position} with value: {value}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to update cell: {str(e)}").dict()
    
    async def _update_cell_by_formula(
        self, 
        result_position: str,
        operator: str,
        start_position: Optional[str] = None,
        end_position: Optional[str] = None,
        position_list: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Update the value of the target cell by applying formulas on specified cells.
        
        Args:
            start_position: The starting position of the range.
            end_position: The ending position of the range.
            position_list: A list of cell positions in A1 notation.
            result_position: The position of the cell for formula result.
            operator: The operator to apply. One of ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'MINUS', 'PRODUCT'].
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        valid_operators = ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'MINUS', 'PRODUCT']
        if operator.upper() not in valid_operators:
            return ToolResult(error=f"Invalid operator: {operator}. Must be one of {valid_operators}").dict()
        
        try:
            # Create cell range reference
            if position_list:
                cell_refs = ",".join(position_list)
            elif start_position and end_position:
                cell_refs = f"{start_position}:{end_position}"
            else:
                return ToolResult(error="Must provide either position_list or start_position and end_position").dict()
            
            # Map to Excel formula syntax
            if operator.upper() == 'MINUS' and len(position_list) == 2:
                formula = f"={position_list[0]}-{position_list[1]}"
            elif operator.upper() == 'PRODUCT' and position_list:
                formula = f"=PRODUCT({cell_refs})"
            else:
                formula = f"={operator.upper()}({cell_refs})"
            
            # Apply formula
            row, col = self._position_to_coordinates(result_position)
            self._current_sheet.cell(row=row, column=col, value=formula)
            
            return ToolResult(output=f"Updated cell {result_position} with formula: {formula}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to update cell by formula: {str(e)}").dict()
    
    async def _update_range(self, start_position: str, end_position: str, values_list: List[List[Any]]) -> Dict[str, Any]:
        """Update a range of cells from a list.
        
        Args:
            start_position: A1 notation of the start cell.
            end_position: A1 notation of the end cell.
            values_list: List of values to be inserted.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            start_row, start_col = self._position_to_coordinates(start_position)
            end_row, end_col = self._position_to_coordinates(end_position)
            
            # Validate range dimensions match values list dimensions
            rows_needed = end_row - start_row + 1
            cols_needed = end_col - start_col + 1
            
            if len(values_list) != rows_needed:
                return ToolResult(error=f"Mismatch in number of rows. Range needs {rows_needed}, got {len(values_list)}").dict()
            
            for row_values in values_list:
                if len(row_values) != cols_needed:
                    return ToolResult(error=f"Mismatch in number of columns. Range needs {cols_needed}").dict()
            
            # Update cells
            for i, row_values in enumerate(values_list):
                for j, value in enumerate(row_values):
                    row = start_row + i
                    col = start_col + j
                    self._current_sheet.cell(row=row, column=col, value=value)
            
            return ToolResult(output=f"Updated range {start_position}:{end_position}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to update range: {str(e)}").dict()
    
    async def _sort_sheet_by_col(self, col_num: int, order: str) -> Dict[str, Any]:
        """Sort the current sheet using the given sort order.
        
        Args:
            col_num: The index of the sort column.
            order: The sort order. Possible values are 'asc' or 'des'.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        if order.lower() not in ['asc', 'des']:
            return ToolResult(error=f"Invalid order: {order}. Must be 'asc' or 'des'").dict()
        
        try:
            # Convert data to list for sorting
            data = list(self._current_sheet.values)
            if not data:
                return ToolResult(output="Sheet is empty, nothing to sort").dict()
            
            # Extract header row if present
            header = data[0]
            data = data[1:]
            
            # Sort data by the specified column
            reverse = order.lower() == 'des'
            data.sort(key=lambda x: x[col_num-1] if col_num-1 < len(x) else None, reverse=reverse)
            
            # Clear sheet
            self._current_sheet.delete_rows(1, self._current_sheet.max_row)
            
            # Reinsert header and sorted data
            self._current_sheet.append(header)
            for row in data:
                self._current_sheet.append(row)
            
            return ToolResult(output=f"Sorted sheet by column {col_num} in {order} order").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to sort sheet: {str(e)}").dict()
    
    async def _merge_cells(self, start_position: str, end_position: str) -> Dict[str, Any]:
        """Merge cells in sheet.
        
        Args:
            start_position: Starting cell position (top left) in A1 annotation.
            end_position: Ending cell position (bottom right) in A1 annotation.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            cell_range = f"{start_position}:{end_position}"
            self._current_sheet.merge_cells(cell_range)
            
            return ToolResult(output=f"Merged cells {cell_range}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to merge cells: {str(e)}").dict()
    
    async def _update_note(self, position: str, content: str) -> Dict[str, Any]:
        """Update a note in a certain cell.
        
        Args:
            position: Cell position in A1 annotation.
            content: The text note to insert.
            
        Returns:
            Result of the operation.
        """
        self._ensure_sheet_selected()
        
        try:
            row, col = self._position_to_coordinates(position)
            cell = self._current_sheet.cell(row=row, column=col)
            cell.comment = content
            
            return ToolResult(output=f"Updated note in cell {position}").dict()
        except Exception as e:
            return ToolResult(error=f"Failed to update note: {str(e)}").dict()
    
    # Data retrieval operations
    async def _get_all_values(self) -> Dict[str, Any]:
        """Display all cell values in current sheet.
        
        Returns:
            All cell values or an error message.
        """
        self._ensure_sheet_selected()
        
        try:
            values = []
            for row in self._current_sheet.iter_rows(values_only=True):
                values.append(list(row))
            
            return ToolResult(output=values).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to get all values: {str(e)}").dict()
    
    async def _get_range_values(self, start_position: str, end_position: str) -> Dict[str, Any]:
        """Returns a list of cell data from a specified range.
        
        Args:
            start_position: Starting cell position in A1 annotation.
            end_position: Ending cell position in A1 annotation.
            
        Returns:
            List of cell data from the specified range or an error message.
        """
        self._ensure_sheet_selected()
        
        try:
            start_row, start_col = self._position_to_coordinates(start_position)
            end_row, end_col = self._position_to_coordinates(end_position)
            
            values = []
            for row in range(start_row, end_row + 1):
                row_values = []
                for col in range(start_col, end_col + 1):
                    cell_value = self._current_sheet.cell(row=row, column=col).value
                    row_values.append(cell_value)
                values.append(row_values)
            
            return ToolResult(output=values).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to get range values: {str(e)}").dict()
    
    async def _get_cell_value(self, position: str) -> Dict[str, Any]:
        """Get the value of a specific cell.
        
        Args:
            position: Cell position in A1 annotation.
            
        Returns:
            Cell value or an error message.
        """
        self._ensure_sheet_selected()
        
        try:
            row, col = self._position_to_coordinates(position)
            value = self._current_sheet.cell(row=row, column=col).value
            
            return ToolResult(output=value).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to get cell value: {str(e)}").dict()
    
    async def _get_value_by_formula(
        self,
        operator: str,
        start_position: Optional[str] = None,
        end_position: Optional[str] = None,
        position_list: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Calculate a value applying formulas on specified cells.
        
        Args:
            start_position: The starting position of the range.
            end_position: The ending position of the range.
            position_list: A list of cell positions in A1 notation.
            operator: The operator to apply.
            
        Returns:
            Calculated result or an error message.
        """
        self._ensure_sheet_selected()
        
        valid_operators = ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'MINUS', 'PRODUCT']
        if operator.upper() not in valid_operators:
            return ToolResult(error=f"Invalid operator: {operator}. Must be one of {valid_operators}").dict()
        
        try:
            # Get values
            values = []
            
            if position_list:
                for pos in position_list:
                    row, col = self._position_to_coordinates(pos)
                    value = self._current_sheet.cell(row=row, column=col).value
                    if value is not None:
                        values.append(value)
            elif start_position and end_position:
                start_row, start_col = self._position_to_coordinates(start_position)
                end_row, end_col = self._position_to_coordinates(end_position)
                
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        value = self._current_sheet.cell(row=row, column=col).value
                        if value is not None:
                            values.append(value)
            else:
                return ToolResult(error="Must provide either position_list or start_position and end_position").dict()
            
            # Apply operator
            if not values:
                return ToolResult(output=0).dict()
            
            if operator.upper() == 'SUM':
                result = sum(values)
            elif operator.upper() == 'AVERAGE':
                result = sum(values) / len(values)
            elif operator.upper() == 'COUNT':
                result = len(values)
            elif operator.upper() == 'MAX':
                result = max(values)
            elif operator.upper() == 'MIN':
                result = min(values)
            elif operator.upper() == 'MINUS' and len(values) == 2:
                result = values[0] - values[1]
            elif operator.upper() == 'PRODUCT':
                result = 1
                for val in values:
                    result *= val
            else:
                return ToolResult(error=f"Cannot apply {operator} to the given values").dict()
            
            return ToolResult(output=result).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to calculate value: {str(e)}").dict()
    
    async def _filter_cells(
        self,
        query: Union[str, re.Pattern],
        in_row: Optional[int] = None,
        in_column: Optional[int] = None
    ) -> Dict[str, Any]:
        """Find all cells matching the query and return their positions.
        
        Args:
            query: A string to match or compiled regular expression.
            in_row: Row number to scope the search.
            in_column: Column number to scope the search.
            
        Returns:
            List of cell addresses that match the query or an error message.
        """
        self._ensure_sheet_selected()
        
        try:
            # Prepare regex pattern if string provided
            if isinstance(query, str):
                pattern = re.compile(query)
            else:
                pattern = query
            
            # Define search area
            min_row = in_row if in_row is not None else 1
            max_row = in_row if in_row is not None else self._current_sheet.max_row
            min_col = in_column if in_column is not None else 1
            max_col = in_column if in_column is not None else self._current_sheet.max_column
            
            # Find matching cells
            matching_cells = []
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    value = self._current_sheet.cell(row=row, column=col).value
                    if value is not None and isinstance(value, str) and pattern.search(value):
                        cell_ref = self._coordinates_to_position(row, col)
                        matching_cells.append(cell_ref)
            
            return ToolResult(output=matching_cells).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to filter cells: {str(e)}").dict()
    
    async def _get_note(self, position: str) -> Dict[str, Any]:
        """Get the note at the certain cell.
        
        Args:
            position: Cell position in A1 annotation.
            
        Returns:
            Note content or an error message.
        """
        self._ensure_sheet_selected()
        
        try:
            row, col = self._position_to_coordinates(position)
            cell = self._current_sheet.cell(row=row, column=col)
            note = cell.comment.text if cell.comment else ""
            
            return ToolResult(output=note).dict()
        except Exception as e:
            return ToolResult(error=f"Failed to get note: {str(e)}").dict()
    
    async def _finish(self) -> Dict[str, Any]:
        """Return all cell values and finish the task.
        
        Returns:
            All cell values or an error message.
        """
        return await self._get_all_values()
